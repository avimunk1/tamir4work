"""Extract data from Excel files according to the defined schema."""

import logging
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

SHEET_PRATIM_KLALIYIM = "פרטים כלליים"
SHEET_TOCHNIT_AVODA = "תוכנית עבודה"


def _safe_cell_value(sheet, cell_ref: str) -> str:
    """Get cell value as string, or empty string if missing/None."""
    try:
        cell = sheet[cell_ref]
        value = cell.value
        return "" if value is None else str(value).strip()
    except (KeyError, AttributeError):
        return ""


def extract_from_file(file_path: Path) -> dict | None:
    """
    Extract smel-mosad and total hours from a single Excel file.

    Returns dict with keys: source, smel_mosad, total_hours.
    Returns None if the file cannot be processed.
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Failed to open %s: %s", file_path.name, e)
        return None

    try:
        source_name = file_path.stem

        smel_mosad = ""
        total_hours = ""

        if SHEET_PRATIM_KLALIYIM in wb.sheetnames:
            sheet = wb[SHEET_PRATIM_KLALIYIM]
            c6 = _safe_cell_value(sheet, "C6")
            c7 = _safe_cell_value(sheet, "C7")
            smel_mosad = f"{c6} {c7}".strip()
        else:
            logger.warning("Sheet '%s' not found in %s", SHEET_PRATIM_KLALIYIM, file_path.name)

        if SHEET_TOCHNIT_AVODA in wb.sheetnames:
            sheet = wb[SHEET_TOCHNIT_AVODA]
            total_hours = _safe_cell_value(sheet, "G21")
        else:
            logger.warning("Sheet '%s' not found in %s", SHEET_TOCHNIT_AVODA, file_path.name)

        return {
            "source": source_name,
            "smel_mosad": smel_mosad,
            "total_hours": total_hours,
        }
    except Exception as e:
        logger.warning("Error extracting from %s: %s", file_path.name, e)
        return None
    finally:
        wb.close()


def extract_from_folder(input_dir: Path) -> list[dict]:
    """
    Extract data from all .xlsx files in the input directory.

    Returns list of dicts with keys: source, smel_mosad, total_hours.
    """
    results = []
    excel_files = sorted(input_dir.glob("*.xlsx"))

    for file_path in excel_files:
        if file_path.name.startswith("~$"):
            continue
        row = extract_from_file(file_path)
        if row is not None:
            results.append(row)
            logger.info("Processed: %s", file_path.name)
        else:
            logger.info("Skipped: %s", file_path.name)

    return results
