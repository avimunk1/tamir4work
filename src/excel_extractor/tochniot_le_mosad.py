"""
Create 'תוכניות למוסד' sheet from 'מוסדות' and 'תוכניות' sheets.

Run as script or as frozen exe. When run without args, opens a file dialog.
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

SHEET_MOSADOT = "מוסדות"
SHEET_TOCHNIOT = "תוכניות"
SHEET_OUTPUT = "תוכניות למוסד"
COLUMNS = ["שם להקמה", "סמל להקמה", "תוכנית", "רשות להקמה"]


def _get_exe_dir() -> Path:
    """When frozen (exe), return exe directory. Otherwise project root."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).resolve().parent.parent.parent


def create_tochniot_le_mosad(file_path: Path) -> int:
    """
    Add 'תוכניות למוסד' sheet to the Excel file.
    Returns number of rows created.
    """
    xlsx = pd.ExcelFile(file_path)

    if SHEET_MOSADOT not in xlsx.sheet_names:
        raise ValueError(f"Sheet '{SHEET_MOSADOT}' not found in file")
    if SHEET_TOCHNIOT not in xlsx.sheet_names:
        raise ValueError(f"Sheet '{SHEET_TOCHNIOT}' not found in file")

    מוסדות = pd.read_excel(xlsx, sheet_name=SHEET_MOSADOT)
    תוכניות = pd.read_excel(xlsx, sheet_name=SHEET_TOCHNIOT)

    rows = []
    for _, מוסד in מוסדות.iterrows():
        for _, תוכנית in תוכניות.iterrows():
            rows.append({
                "שם להקמה": מוסד["שם להקמה"],
                "סמל להקמה": מוסד["סמל להקמה"],
                "תוכנית": תוכנית["תוכנית"],
                "רשות להקמה": מוסד["רשות להקמה"],
            })

    wb = load_workbook(file_path)

    if SHEET_OUTPUT in wb.sheetnames:
        del wb[SHEET_OUTPUT]

    ws = wb.create_sheet(SHEET_OUTPUT, 0)

    for col, header in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row["שם להקמה"])
        ws.cell(row=row_idx, column=2, value=row["סמל להקמה"])
        ws.cell(row=row_idx, column=3, value=row["תוכנית"])
        ws.cell(row=row_idx, column=4, value=row["רשות להקמה"])

    wb.save(file_path)
    return len(rows)


def main() -> None:
    file_path: Path | None = None

    if len(sys.argv) >= 2:
        file_path = Path(sys.argv[1])
        if not file_path.exists():
            _show_error(f"File not found: {file_path}")
            sys.exit(1)
    else:
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            file_path = Path(
                filedialog.askopenfilename(
                    title="Select Excel file (מוסדות + תוכניות)",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    initialdir=_get_exe_dir(),
                )
            )
            root.destroy()
        except Exception as e:
            _show_error(f"Could not open file dialog: {e}\n\nUsage: tochniot_le_mosad.exe <path_to_file.xlsx>")
            sys.exit(1)

        if not file_path or not str(file_path).strip():
            sys.exit(0)

        if not file_path.exists():
            _show_error(f"File not found: {file_path}")
            sys.exit(1)

    try:
        count = create_tochniot_le_mosad(file_path)
        _show_success(f"Created sheet '{SHEET_OUTPUT}' with {count} rows.\n\nSaved to:\n{file_path}")
    except Exception as e:
        _show_error(str(e))
        sys.exit(1)


def _show_success(msg: str) -> None:
    if sys.platform == "win32":
        try:
            import ctypes
            ctypes.windll.user32.MessageBoxW(0, msg, "Success", 0x40)
            return
        except Exception:
            pass
    print(msg)


def _show_error(msg: str) -> None:
    if sys.platform == "win32":
        try:
            import ctypes
            ctypes.windll.user32.MessageBoxW(0, msg, "Error", 0x10)
            return
        except Exception:
            pass
    print(msg, file=sys.stderr)


if __name__ == "__main__":
    main()
